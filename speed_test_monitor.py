import speedtest
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook


# ANSI escape codes for colors
class Colors:
    RESET = "\033[0m"
    BOLD = "\033[1m"
    RED = "\033[91m"
    GREEN = "\033[92m"
    YELLOW = "\033[93m"
    BLUE = "\033[94m"
    WHITE = "\033[97m"


def get_date_time():
    return datetime.now().strftime("%d-%m-%y %H:%M:%S")


def run_speed_test():
    retry_count = 3

    for _ in range(retry_count):
        try:
            print(f"{Colors.WHITE}{get_date_time()}: {Colors.GREEN}{Colors.BOLD}Testing--------{Colors.RESET}")
            st = speedtest.Speedtest()
            download_speed = st.download() / 1024 / 1024  # Convert to Mbps
            upload_speed = st.upload() / 1024 / 1024  # Convert to Mbps
            print(
                f"{get_date_time()}: {Colors.GREEN}{Colors.BOLD}Success--------{Colors.RESET} Download Speed: {Colors.BLUE}{download_speed:.2f} Mbps{Colors.RESET}, Upload Speed: {Colors.BLUE}{upload_speed:.2f} Mbps{Colors.RESET}")
            return download_speed, upload_speed, "Success", ""
        except Exception as e:
            print(f"{get_date_time()}: {Colors.RED}{Colors.BOLD}FAILED: {e}{Colors.RESET}")
            return 0, 0, "Failed", str(e)
        finally:
            time.sleep(5)  # Wait for 5 seconds before retrying


def write_to_excel(data, filename="speed_test_results.xlsx"):
    try:
        # Load existing workbook or create a new one
        try:
            workbook = load_workbook(filename)
        except FileNotFoundError:
            workbook = Workbook()

        # Select the active sheet
        sheet = workbook.active

        # Write headers if the sheet is empty
        if sheet.max_row == 1:
            headers = ["Test Number", "Time", "Status", "Error", "Download Speed (Mbps)", "Upload Speed (Mbps)"]
            sheet.append(headers)

        # Write data as a new row
        sheet.append(data)

        # Save the workbook
        workbook.save(filename)
    except Exception as e:
        print(f"{get_date_time()}: {Colors.RED}Failed to write to Excel file: {e}{Colors.RESET}")


def main(interval_minutes=15, num_tests=96):
    for test_number in range(1, num_tests + 1):
        try:
            download_speed, upload_speed, status, error = run_speed_test()
            data = [test_number, get_date_time(), status, error, download_speed, upload_speed]
            write_to_excel(data)
        except Exception as e:
            print(f"{get_date_time()}: {Colors.RED}ERROR: {e}{Colors.RESET}")
        print(f"{get_date_time()}: {Colors.WHITE}Going back to sleep{Colors.RESET}")
        time.sleep(interval_minutes * 60)  # Convert minutes to seconds


if __name__ == "__main__":
    main()
